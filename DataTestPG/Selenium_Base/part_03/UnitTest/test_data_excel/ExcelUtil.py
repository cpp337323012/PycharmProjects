# encoding:utf-8

from openpyxl import load_workbook


class ParaseExcel:

    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        dataList = []
        for line in self.sheet:
            tmpList = []
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            dataList.append(tmpList)
        return dataList[1:]

if __name__ == '__main__':
    excelPath = './测试数据.xlsx'
    sheetName = '工作表1'
    pe = ParaseExcel(excelPath, sheetName)
    for i in pe.getDatasFromSheet():
        print(i[0], i[1])